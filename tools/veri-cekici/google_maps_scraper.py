"""
Google Maps İşyeri Veri Çekici
Gereksinimler:
    pip install playwright openpyxl requests beautifulsoup4
    playwright install chromium
"""

import threading
import time
import re
import os
from datetime import datetime
import json

# tkinter masaustu GUI icindir (BASLAT.bat ile calistirilir); sunucu tarafinda
# (PHP -> bu dosyadan sadece scrape_google_maps cagirilir) genelde kurulu
# olmaz ve headless Linux'ta modul import'unu kirar. Bu yuzden opsiyonel.
try:
    import tkinter as tk
    from tkinter import ttk, messagebox, filedialog
    TK_OK = True
except ImportError:
    TK_OK = False

# Kütüphane kontrol
try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_OK = True
except ImportError:
    PLAYWRIGHT_OK = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    import requests
    from bs4 import BeautifulSoup
    REQUESTS_OK = True
except ImportError:
    REQUESTS_OK = False


# ─────────────────────────────────────────────
#  SCRAPER FONKSİYONLARI
# ─────────────────────────────────────────────

def find_email_on_website(url: str, timeout: int = 8) -> str:
    """Web sitesini ziyaret edip e-posta adresi bulmaya çalışır."""
    if not REQUESTS_OK or not url:
        return ""
    try:
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        r = requests.get(url, headers=headers, timeout=timeout, allow_redirects=True)
        text = r.text

        # Sayfada e-posta ara
        emails = re.findall(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", text)
        blacklist = {"example.com", "yourdomain.com", "domain.com", "sentry.io",
                     "email.com", "test.com", "wixpress.com", "squarespace.com"}
        for email in emails:
            domain = email.split("@")[-1].lower()
            if domain not in blacklist and not email.startswith("noreply"):
                return email.lower()

        # İletişim sayfasını dene
        soup = BeautifulSoup(text, "html.parser")
        contact_links = []
        for a in soup.find_all("a", href=True):
            href = a["href"].lower()
            if any(k in href for k in ["iletisim", "contact", "about", "hakkimizda"]):
                contact_links.append(a["href"])

        for link in contact_links[:2]:
            try:
                full = link if link.startswith("http") else url.rstrip("/") + "/" + link.lstrip("/")
                r2 = requests.get(full, headers=headers, timeout=timeout)
                emails2 = re.findall(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", r2.text)
                for email in emails2:
                    domain = email.split("@")[-1].lower()
                    if domain not in blacklist and not email.startswith("noreply"):
                        return email.lower()
            except Exception:
                pass
    except Exception:
        pass
    return ""


def scrape_google_maps(query: str, max_results: int, log_fn, stop_event,
                       find_email: bool = True) -> list[dict]:
    """Google Maps'ten işyeri verisi çeker."""
    if not PLAYWRIGHT_OK:
        raise RuntimeError(
            "playwright kurulu degil. Sunucuda calistirmak icin: "
            "pip install playwright && playwright install chromium --with-deps"
        )

    results = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=["--no-sandbox"])
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            viewport={"width": 1280, "height": 800}
        )
        page = context.new_page()

        try:
            search_url = f"https://www.google.com/maps/search/{query.replace(' ', '+')}"
            log_fn(f"🔍 Aranıyor: {query}")
            page.goto(search_url, wait_until="domcontentloaded", timeout=30000)
            time.sleep(3)

            # Çerez kabul
            try:
                accept = page.query_selector('button:has-text("Accept"), button:has-text("Kabul")')
                if accept:
                    accept.click()
                    time.sleep(1)
            except Exception:
                pass

            # ── Scroll: yeterli sonuç yüklenene kadar listeyi kaydır ──
            FEED = '[role="feed"]'
            ITEM_SEL = '[role="feed"] > div > div > a'

            scrollable = page.query_selector(FEED)
            if not scrollable:
                # Alternatif seçici dene
                scrollable = page.query_selector('div[aria-label]')

            log_fn("📜 Sonuçlar yükleniyor...")
            prev_count = 0
            no_change = 0
            while True:
                if stop_event.is_set():
                    break
                count = page.evaluate('(sel) => document.querySelectorAll(sel).length', ITEM_SEL)
                if count >= max_results:
                    break
                if count == prev_count:
                    no_change += 1
                    if no_change >= 5:
                        break
                else:
                    no_change = 0
                prev_count = count
                # JavaScript ile scroll — element referansı bayatlamaz
                page.evaluate("""
                    const feed = document.querySelector('[role="feed"]');
                    if (feed) feed.scrollBy(0, 1200);
                    else window.scrollBy(0, 1200);
                """)
                time.sleep(1.5)

            total = page.evaluate('(sel) => document.querySelectorAll(sel).length', ITEM_SEL)
            log_fn(f"✅ {total} yer bulundu, detaylar çekiliyor...")

            # ── Her işyeri için URL'yi al, sayfaya git, veri çek ──
            # Önce tüm href'leri topla (element referansı değil, string)
            hrefs = page.evaluate(
                '([sel, max]) => Array.from(document.querySelectorAll(sel)).slice(0, max).map(a => a.href)',
                [ITEM_SEL, max_results]
            )

            def get_text(sel):
                el = page.query_selector(sel)
                return el.inner_text().strip() if el else ""

            for i, href in enumerate(hrefs):
                if stop_event.is_set():
                    log_fn("⛔ Kullanıcı durdurdu.")
                    break

                try:
                    # Direkt URL'ye git — stale element sorunu yok
                    page.goto(href, wait_until="domcontentloaded", timeout=20000)
                    time.sleep(2)

                    # Koordinat: sonuc listesinden dogrudan href ile bir yer
                    # sayfasina gidince URL "/@enlem,boylam,zoom" DEGIL,
                    # "...!3d<enlem>!4d<boylam>!..." (data= parametresi
                    # icinde, Google'in kendi kisa-kod formati) seklinde
                    # geliyor. Once bunu dene, olmazsa eski /@ formatina
                    # (kullanici haritada gezinirken olusan URL'ler icin) dus.
                    lat, lng = "", ""
                    coord_match = re.search(r'!3d(-?\d+\.\d+)!4d(-?\d+\.\d+)', page.url)
                    if not coord_match:
                        coord_match = re.search(r'/@(-?\d+\.\d+),(-?\d+\.\d+)', page.url)
                    if coord_match:
                        lat, lng = coord_match.group(1), coord_match.group(2)

                    name = get_text('h1.DUwDvf') or get_text('h1')

                    # Adres — birden fazla seçici dene
                    address = ""
                    for sel in ['button[data-item-id^="address"]',
                                'button[data-tooltip="Adresi kopyala"]',
                                '[data-item-id*="address"]']:
                        el = page.query_selector(sel)
                        if el:
                            address = el.get_attribute("aria-label") or el.inner_text()
                            address = re.sub(r'^(Adres|Address):\s*', '', address).strip()
                            break

                    # Telefon
                    phone = ""
                    for sel in ['button[data-item-id^="phone"]',
                                'button[data-tooltip="Telefon numarasını kopyala"]',
                                '[data-item-id*="phone"]']:
                        el = page.query_selector(sel)
                        if el:
                            phone = el.get_attribute("aria-label") or el.inner_text()
                            phone = re.sub(r'^(Telefon|Phone):\s*', '', phone).strip()
                            break

                    # Web sitesi
                    website = ""
                    web_a = page.query_selector('a[data-item-id^="authority"]')
                    if web_a:
                        website = web_a.get_attribute("href") or ""

                    # Puan & Kategori
                    rating = get_text('div.F7nice span[aria-hidden="true"]')
                    category = get_text('button.DkEaL') or get_text('[jsaction*="category"]')

                    # E-posta
                    email = ""
                    if find_email and website:
                        log_fn(f"  📧 E-posta aranıyor: {name[:30]}...")
                        email = find_email_on_website(website)

                    if name:
                        results.append({
                            "İşyeri Adı": name,
                            "Kategori": category,
                            "Adres": address,
                            "Telefon": phone,
                            "Web Sitesi": website,
                            "E-posta": email,
                            "Puan": rating,
                            "Enlem": lat,
                            "Boylam": lng,
                        })
                        log_fn(f"  [{i+1}] ✓ {name} | {phone} | {email or '-'}")
                    else:
                        log_fn(f"  [{i+1}] ⚠️ İsim alınamadı, atlandı.")

                except Exception as e:
                    log_fn(f"  [{i+1}] ⚠️ Hata: {str(e)[:120]}")

        finally:
            browser.close()

    return results


def export_to_excel(data: list[dict], path: str) -> None:
    """Veriyi Excel dosyasına yazar."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Google Maps Verileri"

    if not data:
        wb.save(path)
        return

    headers = list(data[0].keys())

    # Başlık stili
    header_fill = PatternFill("solid", fgColor="1A73E8")
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Başlıklar
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    ws.row_dimensions[1].height = 30

    # Veri satırları
    alt_fill = PatternFill("solid", fgColor="EAF2FF")
    for row_idx, record in enumerate(data, 2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(key, ""))
            cell.fill = fill
            cell.alignment = left
            cell.border = border
            cell.font = Font(name="Arial", size=10)

    # Sütun genişlikleri
    widths = {"İşyeri Adı": 35, "Kategori": 20, "Adres": 45,
              "Telefon": 18, "Web Sitesi": 35, "E-posta": 30, "Puan": 8,
              "Enlem": 14, "Boylam": 14}
    for col, key in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = widths.get(key, 20)

    # Filtre ekle
    ws.auto_filter.ref = ws.dimensions

    # Özet sayfası
    ws2 = wb.create_sheet("Özet")
    ws2["A1"] = "Arama Özeti"
    ws2["A1"].font = Font(bold=True, size=14, name="Arial")
    ws2["A2"] = f"Toplam kayıt:"
    ws2["B2"] = len(data)
    ws2["A3"] = f"Oluşturma tarihi:"
    ws2["B3"] = datetime.now().strftime("%d.%m.%Y %H:%M")
    ws2["A4"] = "E-posta bulunan:"
    ws2["B4"] = sum(1 for r in data if r.get("E-posta"))
    ws2["A5"] = "Telefon bulunan:"
    ws2["B5"] = sum(1 for r in data if r.get("Telefon"))

    for cell in ["A2", "A3", "A4", "A5"]:
        ws2[cell].font = Font(bold=True, name="Arial")

    wb.save(path)


# ─────────────────────────────────────────────
#  TKINTER GUI
# ─────────────────────────────────────────────

# tkinter kurulu degilse App sinifi normal object'ten turer; sinif govdesindeki
# tk./ttk. referanslari metot icinde oldugu icin (lazy), bu sadece App()
# gercekten olusturulmaya calisilirsa (bkz. __main__) hataya duser, modul
# import edilirken degil.
_TkBase = tk.Tk if TK_OK else object


class App(_TkBase):
    def __init__(self):
        super().__init__()
        self.title("🗺️ Google Maps İşyeri Veri Çekici")
        self.geometry("820x680")
        self.resizable(True, True)
        self.configure(bg="#F0F4F8")
        self._check_deps()
        self._build_ui()
        self.results: list[dict] = []
        self.stop_event = threading.Event()

    def _check_deps(self):
        missing = []
        if not PLAYWRIGHT_OK:
            missing.append("playwright  →  pip install playwright && playwright install chromium")
        if not OPENPYXL_OK:
            missing.append("openpyxl   →  pip install openpyxl")
        if not REQUESTS_OK:
            missing.append("requests / beautifulsoup4  →  pip install requests beautifulsoup4")
        if missing:
            msg = "Eksik kütüphaneler:\n\n" + "\n".join(missing)
            messagebox.showwarning("Kurulum Gerekli", msg)

    def _build_ui(self):
        # Başlık
        header = tk.Frame(self, bg="#1A73E8", pady=12)
        header.pack(fill="x")
        tk.Label(header, text="🗺️  Google Maps İşyeri Veri Çekici",
                 font=("Arial", 16, "bold"), fg="white", bg="#1A73E8").pack()
        tk.Label(header, text="Sektör · İl · İlçe bazlı arama — Excel'e aktar",
                 font=("Arial", 10), fg="#C8DCFF", bg="#1A73E8").pack()

        # Form alanı
        form = tk.LabelFrame(self, text=" Arama Parametreleri ", font=("Arial", 10, "bold"),
                             bg="#F0F4F8", fg="#333", padx=16, pady=12)
        form.pack(fill="x", padx=16, pady=(12, 6))

        fields = [
            ("Sektör / İşyeri Türü:", "sektor", "Örn: restoran, eczane, dişçi, avukat"),
            ("İl:", "il", "Örn: Bursa, İstanbul, Ankara"),
            ("İlçe (opsiyonel):", "ilce", "Örn: Nilüfer, Osmangazi (boş bırakılabilir)"),
        ]
        self.vars = {}
        for label, key, ph in fields:
            row = tk.Frame(form, bg="#F0F4F8")
            row.pack(fill="x", pady=4)
            tk.Label(row, text=label, width=22, anchor="w",
                     font=("Arial", 10), bg="#F0F4F8").pack(side="left")
            var = tk.StringVar()
            entry = ttk.Entry(row, textvariable=var, font=("Arial", 10), width=38)
            entry.pack(side="left", padx=4)
            tk.Label(row, text=ph, font=("Arial", 9), fg="#999", bg="#F0F4F8").pack(side="left")
            self.vars[key] = var

        # Ayarlar
        opt = tk.Frame(form, bg="#F0F4F8")
        opt.pack(fill="x", pady=(8, 2))
        tk.Label(opt, text="Maksimum sonuç:", width=22, anchor="w",
                 font=("Arial", 10), bg="#F0F4F8").pack(side="left")
        self.max_var = tk.IntVar(value=20)
        ttk.Spinbox(opt, from_=5, to=1000, increment=5, textvariable=self.max_var,
                    width=6, font=("Arial", 10)).pack(side="left")

        self.email_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(opt, text="E-posta ara (web sitesini tarar, daha yavaş)",
                        variable=self.email_var).pack(side="left", padx=20)

        # Butonlar
        btn_row = tk.Frame(self, bg="#F0F4F8")
        btn_row.pack(pady=8)

        self.btn_start = tk.Button(btn_row, text="▶  Aramayı Başlat",
                                   command=self._start, font=("Arial", 11, "bold"),
                                   bg="#1A73E8", fg="white", relief="flat",
                                   padx=20, pady=8, cursor="hand2",
                                   activebackground="#155DB2")
        self.btn_start.pack(side="left", padx=6)

        self.btn_stop = tk.Button(btn_row, text="⏹  Durdur",
                                  command=self._stop, font=("Arial", 11),
                                  bg="#EA4335", fg="white", relief="flat",
                                  padx=20, pady=8, cursor="hand2",
                                  state="disabled", activebackground="#C5221F")
        self.btn_stop.pack(side="left", padx=6)

        self.btn_excel = tk.Button(btn_row, text="💾  Excel'e Kaydet",
                                   command=self._save_excel, font=("Arial", 11),
                                   bg="#34A853", fg="white", relief="flat",
                                   padx=20, pady=8, cursor="hand2",
                                   state="disabled", activebackground="#2D8E47")
        self.btn_excel.pack(side="left", padx=6)

        # İlerleme
        self.progress = ttk.Progressbar(self, mode="indeterminate", length=400)
        self.progress.pack(pady=(0, 4))

        self.status_var = tk.StringVar(value="Hazır")
        tk.Label(self, textvariable=self.status_var, font=("Arial", 9),
                 fg="#555", bg="#F0F4F8").pack()

        # Log kutusu
        log_frame = tk.LabelFrame(self, text=" İşlem Kaydı ", font=("Arial", 10, "bold"),
                                   bg="#F0F4F8", fg="#333")
        log_frame.pack(fill="both", expand=True, padx=16, pady=(6, 8))
        self.log_box = tk.Text(log_frame, font=("Consolas", 9), bg="#1E1E1E",
                               fg="#D4D4D4", relief="flat", state="disabled",
                               wrap="word")
        scroll = ttk.Scrollbar(log_frame, command=self.log_box.yview)
        self.log_box.configure(yscrollcommand=scroll.set)
        scroll.pack(side="right", fill="y")
        self.log_box.pack(fill="both", expand=True, padx=2, pady=2)

        # Sonuç sayacı
        self.count_var = tk.StringVar(value="Sonuç: 0")
        tk.Label(self, textvariable=self.count_var, font=("Arial", 10, "bold"),
                 fg="#1A73E8", bg="#F0F4F8").pack(pady=(0, 4))

    def log(self, msg: str):
        self.after(0, self._log_main, msg)

    def _log_main(self, msg: str):
        self.log_box.configure(state="normal")
        self.log_box.insert("end", msg + "\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def _start(self):
        sektor = self.vars["sektor"].get().strip()
        il = self.vars["il"].get().strip()
        ilce = self.vars["ilce"].get().strip()

        if not sektor or not il:
            messagebox.showwarning("Eksik Bilgi", "Sektör ve İl alanlarını doldurun.")
            return
        if not PLAYWRIGHT_OK:
            messagebox.showerror("Eksik Kütüphane",
                                 "Playwright kurulu değil.\nTerminalde:\n"
                                 "pip install playwright\nplaywright install chromium")
            return

        query_parts = [sektor, ilce, il] if ilce else [sektor, il]
        query = " ".join(query_parts)

        self.results = []
        self.stop_event.clear()
        self.btn_start.config(state="disabled")
        self.btn_stop.config(state="normal")
        self.btn_excel.config(state="disabled")
        self.progress.start(12)
        self.status_var.set("Aranıyor...")
        self.count_var.set("Sonuç: 0")
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

        threading.Thread(target=self._run_scrape,
                         args=(query, self.max_var.get(), self.email_var.get()),
                         daemon=True).start()

    def _run_scrape(self, query: str, max_r: int, find_email: bool):
        try:
            data = scrape_google_maps(query, max_r, self.log, self.stop_event, find_email)
            self.results = data
            self.after(0, self._done, len(data))
        except Exception as e:
            self.after(0, self._error, str(e))

    def _done(self, n: int):
        self.progress.stop()
        self.btn_start.config(state="normal")
        self.btn_stop.config(state="disabled")
        self.btn_excel.config(state="normal" if self.results else "disabled")
        self.status_var.set(f"Tamamlandı — {n} kayıt bulundu.")
        self.count_var.set(f"Sonuç: {n}")
        self.log(f"\n🎉 Tamamlandı! {n} işyeri verisi çekildi.")
        self.log("💾 'Excel'e Kaydet' butonuna tıklayarak dışa aktarabilirsiniz.")

    def _error(self, msg: str):
        self.progress.stop()
        self.btn_start.config(state="normal")
        self.btn_stop.config(state="disabled")
        self.status_var.set("Hata oluştu!")
        self.log(f"❌ HATA: {msg}")
        messagebox.showerror("Hata", msg)

    def _stop(self):
        self.stop_event.set()
        self.log("⏹ Durdurma isteği gönderildi...")
        self.btn_stop.config(state="disabled")

    def _save_excel(self):
        if not self.results:
            messagebox.showinfo("Veri Yok", "Kaydedilecek veri yok.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyası", "*.xlsx")],
            initialfile=f"google_maps_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )
        if not path:
            return
        try:
            export_to_excel(self.results, path)
            self.log(f"✅ Excel kaydedildi: {path}")
            messagebox.showinfo("Başarılı", f"{len(self.results)} kayıt Excel'e kaydedildi!\n\n{path}")
        except Exception as e:
            messagebox.showerror("Hata", f"Excel kaydedilemedi:\n{e}")


# ─────────────────────────────────────────────
if __name__ == "__main__":
    if not TK_OK:
        print("HATA: tkinter kurulu degil. Bu masaustu programi GUI icin tkinter gerektirir.")
        print("Windows/Mac Python kurulumlarinda genelde hazir gelir; Linux'ta 'sudo apt install python3-tk' ile kurulabilir.")
        raise SystemExit(1)

    app = App()
    app.mainloop()
